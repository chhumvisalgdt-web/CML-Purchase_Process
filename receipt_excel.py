"""Excel layer for goods receipt: generate the outstanding-lines file, read it
back, write the validation report. openpyxl only -- no Telegram, no gspread.

Reader contract (must stay in sync with build_receipt_file):
    SHEET_NAME = "Receiving"
    C3 = invoice number, C4 = invoice date   (both mandatory)
    DATA_FIRST = 8 ..                        (grows with the line count)
    B = CML code (locked)   F = receiving now   G = invoice qty
    H = lot no.             I = expiry          J = note
Columns C/D/E (item, pack, ordered/already received) are locked context. The
reader never trusts them: line_id, ordered and received-to-date come from the
PO and the Receipts tab, not from the file.
"""
import re
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import receipt_validate as rv

SHEET_NAME = "Receiving"
META_SHEET = "_meta"
DATA_FIRST = 8
ROWS_PER_LINE = 2          # a line can arrive as two lots without adding rows
SPARE_ROWS = 4             # ... and more lots are still possible

COL_CODE, COL_ITEM, COL_PACK, COL_CTX = 2, 3, 4, 5
COL_QTY, COL_INVQTY, COL_LOT, COL_EXP, COL_NOTE = 6, 7, 8, 9, 10

FONT = "Arial"
INK = "1F1F1F"
MUTED = "5F5E5A"
BLUE = "0000FF"
HEAD_FILL = PatternFill("solid", fgColor="187B85")
INPUT_FILL = PatternFill("solid", fgColor="FFF6D9")
LOCK_FILL = PatternFill("solid", fgColor="F4F3EF")
OK_FILL = PatternFill("solid", fgColor="EAF3DE")
BAD_FILL = PatternFill("solid", fgColor="FCEBEB")
WARN_FILL = PatternFill("solid", fgColor="FDF1DC")
HAIR = Side(style="thin", color="C9C7BF")
BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

HEADERS = ["No.", "CML Code", "Item", "Pack", "Ordered / received so far",
           "Receiving now", "Invoice qty", "Lot no.", "Expiry", "Note"]
WIDTHS = [6, 16, 42, 16, 22, 14, 12, 18, 14, 30]


def _slug(t):
    return re.sub(r"[^A-Za-z0-9]+", "_", str(t)).strip("_")[:40] or "po"


def build_receipt_file(po_no, outstanding, out=None):
    """One file per delivery, holding only the lines not yet fully received.

    Each line gets ROWS_PER_LINE blank rows so a split-lot delivery needs no
    improvisation, plus SPARE_ROWS at the end for a third or fourth lot.
    """
    lines = sorted(outstanding.items(), key=lambda kv: kv[1]["line_id"])
    if not lines:
        raise ValueError(f"PO #{po_no} has no outstanding lines")

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"Goods receipt - PO #{po_no}"
    t.font = Font(name=FONT, size=14, bold=True, color="187B85")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:J2")
    ws["A2"] = ("Enter what you actually counted, not what the invoice says. "
                "One row per lot - a line arriving as two lots uses both its rows.")
    ws["A2"].font = Font(name=FONT, size=9, color=MUTED)

    for ref, label in (("B3", "Invoice no."), ("B4", "Invoice date")):
        ws[ref] = label
        ws[ref].font = Font(name=FONT, size=10, bold=True, color=MUTED)
    for ref in ("C3", "C4"):
        ws[ref].fill = INPUT_FILL
        ws[ref].border = BOX
        ws[ref].protection = Protection(locked=False)
        ws[ref].font = Font(name=FONT, size=10, color=BLUE)
    ws["C4"].number_format = "dd-mmm-yyyy"
    ws["D3"] = "required"
    ws["D4"] = "required"
    for ref in ("D3", "D4"):
        ws[ref].font = Font(name=FONT, size=9, italic=True, color=MUTED)

    for i, (head, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=7, column=i, value=head)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BOX
    ws.row_dimensions[7].height = 26

    r = DATA_FIRST
    for n, (code, line) in enumerate(lines, start=1):
        for slot in range(ROWS_PER_LINE):
            ws.row_dimensions[r].height = 17
            first = slot == 0
            no = ws.cell(row=r, column=1, value=n if first else "")
            no.font = Font(name=FONT, size=10, color=MUTED)
            no.alignment = Alignment(horizontal="center")
            no.border = BOX

            ctx = (f"{line['ordered']} ordered"
                   + (f", {line['received']} received"
                      if line["received"] else ""))
            for col, val in ((COL_CODE, code),
                             (COL_ITEM, line["item"] if first else ""),
                             (COL_PACK, line["pack"] if first else ""),
                             (COL_CTX, ctx if first else "")):
                c = ws.cell(row=r, column=col, value=val)
                c.font = Font(name=FONT, size=10,
                              color=INK if first else MUTED)
                c.fill = LOCK_FILL
                c.border = BOX
                if col == COL_CODE:
                    c.number_format = "@"

            for col in (COL_QTY, COL_INVQTY, COL_LOT, COL_EXP, COL_NOTE):
                c = ws.cell(row=r, column=col)
                c.font = Font(name=FONT, size=10, color=BLUE)
                c.fill = INPUT_FILL
                c.border = BOX
                c.protection = Protection(locked=False)
            ws.cell(row=r, column=COL_QTY).number_format = "0"
            ws.cell(row=r, column=COL_INVQTY).number_format = "0"
            ws.cell(row=r, column=COL_EXP).number_format = "dd-mmm-yyyy"
            r += 1

    last_line_row = r - 1
    for _ in range(SPARE_ROWS):
        ws.row_dimensions[r].height = 17
        for col in (COL_CODE, COL_QTY, COL_INVQTY, COL_LOT, COL_EXP, COL_NOTE):
            c = ws.cell(row=r, column=col)
            c.font = Font(name=FONT, size=10, color=BLUE)
            c.fill = INPUT_FILL
            c.border = BOX
            c.protection = Protection(locked=False)
            if col == COL_CODE:
                c.number_format = "@"
        ws.cell(row=r, column=COL_QTY).number_format = "0"
        ws.cell(row=r, column=COL_EXP).number_format = "dd-mmm-yyyy"
        r += 1
    data_last = r - 1

    ws.merge_cells(f"A{data_last + 1}:J{data_last + 1}")
    g = ws[f"A{data_last + 1}"]
    g.value = (f"Spare rows above are for extra lots - copy the CML code into "
               f"column B. Only rows {DATA_FIRST}-{data_last} are read.")
    g.font = Font(name=FONT, size=9, italic=True, color="A32D2D")

    dv_code = DataValidation(
        type="list", formula1='"' + ",".join(c for c, _ in lines)[:250] + '"',
        allow_blank=True, showErrorMessage=False)
    if len(",".join(c for c, _ in lines)) <= 250:
        ws.add_data_validation(dv_code)
        dv_code.add(f"B{last_line_row + 1}:B{data_last}")

    dv_qty = DataValidation(
        type="whole", operator="between", formula1=-9999, formula2=9999,
        allow_blank=True, showErrorMessage=True, errorTitle="Invalid quantity",
        error="Quantity must be a whole number. Negative values are corrections.")
    ws.add_data_validation(dv_qty)
    dv_qty.add(f"F{DATA_FIRST}:G{data_last}")

    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.freeze_panes = f"A{DATA_FIRST}"

    mw = wb.create_sheet(META_SHEET)
    for i, (k, v) in enumerate([("kind", "receipt"), ("po_no", po_no),
                                ("data_first", DATA_FIRST),
                                ("data_last", data_last),
                                ("generated_at",
                                 datetime.now().isoformat(timespec="seconds"))],
                               start=1):
        mw.cell(row=i, column=1, value=k)
        mw.cell(row=i, column=2, value=str(v))
    mw.sheet_state = "hidden"

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if out:
        open(out, "wb").write(data)
    return {"bytes": data, "filename": f"Receipt_PO_{_slug(po_no)}.xlsx",
            "lines": len(lines), "data_last": data_last}


def read_receipt(data):
    """Returns (rows, meta). meta carries workbook-level facts captured before
    row iteration, so a dropped row is detectable from the record alone."""
    src = data if isinstance(data, (bytes, bytearray)) else data.getvalue()
    meta = {"kind": "", "po_no": "", "sheet_count": 0, "sheet_names": "",
            "max_row": 0, "rows_in_range": 0, "populated_below_range": 0,
            "invoice_no": "", "invoice_date": None, "error": ""}
    try:
        wb = load_workbook(BytesIO(src), data_only=True)
    except Exception as exc:
        meta["error"] = f"not a readable Excel file ({type(exc).__name__})"
        return [], meta

    meta["sheet_count"] = len(wb.sheetnames)
    meta["sheet_names"] = ", ".join(wb.sheetnames)

    data_last = None
    if META_SHEET in wb.sheetnames:
        mw = wb[META_SHEET]
        got = {str(mw.cell(row=i, column=1).value or ""):
               str(mw.cell(row=i, column=2).value or "")
               for i in range(1, mw.max_row + 1)}
        meta["kind"] = got.get("kind", "")
        meta["po_no"] = got.get("po_no", "")
        try:
            data_last = int(got.get("data_last", ""))
        except ValueError:
            data_last = None

    if SHEET_NAME not in wb.sheetnames:
        meta["error"] = (f"no '{SHEET_NAME}' tab in this file. Send /receive "
                         f"<po number> for a fresh one.")
        return [], meta

    ws = wb[SHEET_NAME]
    meta["max_row"] = ws.max_row or 0
    meta["invoice_no"] = rv.norm_text(ws["C3"].value)
    meta["invoice_date"] = ws["C4"].value
    if data_last is None:
        data_last = meta["max_row"]

    rows = []
    for r in range(DATA_FIRST, data_last + 1):
        rows.append({
            "row_no": r,
            "code": ws.cell(row=r, column=COL_CODE).value,
            "qty": ws.cell(row=r, column=COL_QTY).value,
            "invoice_qty": ws.cell(row=r, column=COL_INVQTY).value,
            "lot": ws.cell(row=r, column=COL_LOT).value,
            "expiry": ws.cell(row=r, column=COL_EXP).value,
            "note": ws.cell(row=r, column=COL_NOTE).value,
        })
    meta["rows_in_range"] = len(rows)

    beyond = 0
    for r in range(data_last + 2, (ws.max_row or 0) + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "")
               for c in (COL_CODE, COL_QTY)):
            beyond += 1
    meta["populated_below_range"] = beyond
    return rows, meta


REPORT_HEADERS = ["Row", "CML Code", "Qty", "Lot no.", "Expiry", "Status",
                  "What to do", "Item"]


def write_report(result, meta=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipt check"
    s = result.summary
    blocked = s.get("rows_blocked", 0)

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = ("Receipt check - all rows passed" if not blocked
               else f"Receipt check - {blocked} row(s) need fixing")
    t.font = Font(name=FONT, size=14, bold=True,
                  color="0F6E56" if not blocked else "A32D2D")

    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Invoice {s.get('invoice_no') or '(missing)'} \u00b7 "
                f"{s.get('rows_populated', 0)} row(s) filled in \u00b7 "
                f"{s.get('rows_ok', 0)} accepted \u00b7 {blocked} blocked. "
                f"Nothing has been recorded.")
    ws["A2"].font = Font(name=FONT, size=10, color=MUTED)

    for i, (h, w) in enumerate(zip(REPORT_HEADERS,
                                   [7, 16, 9, 18, 14, 18, 56, 40]), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.border = BOX

    for i, r in enumerate(result.report, start=5):
        st = r["status"]
        fill = (OK_FILL if st == rv.ST_OK else
                WARN_FILL if st in rv.CONFIRMABLE else BAD_FILL)
        vals = [r["row_no"], r["raw_code"], r["raw_qty"], r["raw_lot"],
                r["raw_expiry"],
                "OK" if st == rv.ST_OK else st.replace("_", " "),
                r["message"], r["matched_item"]]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name=FONT, size=10, color=INK)
            c.fill = fill
            c.border = BOX
            c.alignment = Alignment(vertical="top", wrap_text=col in (7, 8))
    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    return {"bytes": buf.getvalue(), "filename": "Receipt_check.xlsx"}
