"""Excel layer for the PO upload path: generate template, read it back, write
the validation report.

Only openpyxl here -- no Telegram, no gspread -- so this is testable against a
temp file. The parser contract lives at the top and the generator honours it.
"""
import hashlib
import re
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from clock import local_now

from upload_validate import (CAT_OTHER, CAT_REAGENT, CATEGORY_LABEL, MAX_LINES,
                             STATUS_OK, norm_code)

TEMPLATE_VERSION = "1.0"
SHEET_NAME = "PO Request"
MASTER_SHEET = "Master"
DATA_FIRST = 8
DATA_LAST = DATA_FIRST + MAX_LINES - 1
COL_CODE, COL_QTY, COL_NOTE = 2, 6, 7
META_CELLS = {"version": "B5", "generated": "D5", "category": "F5",
              "supplier": "G5"}

FONT = "Arial"
INK = "1F1F1F"
MUTED = "5F5E5A"
INPUT_BLUE = "0000FF"
HEAD_FILL = PatternFill("solid", fgColor="187B85")
INPUT_FILL = PatternFill("solid", fgColor="FFF6D9")
LOOKUP_FILL = PatternFill("solid", fgColor="F4F3EF")
BAD_FILL = PatternFill("solid", fgColor="FCEBEB")
OK_FILL = PatternFill("solid", fgColor="EAF3DE")
HAIR = Side(style="thin", color="C9C7BF")
BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

HEADERS = ["No.", "Material Code", "Item (auto)", "Supplier (auto)",
           "Pack (auto)", "Qty", "Note"]
WIDTHS = [6, 20, 44, 26, 16, 9, 34]


def _lookup(code_cell, master_col, n_master):
    out = f"{MASTER_SHEET}!${master_col}$2:${master_col}${n_master + 1}"
    key = f"{MASTER_SHEET}!$A$2:$A${n_master + 1}"
    return (f'=IF({code_cell}="","",'
            f'IFERROR(INDEX({out},MATCH({code_cell},{key},0)),"CODE NOT FOUND"))')


def build_template(index, category, supplier, out_path=None):
    """Rows come from MasterIndex.rows_for(), which already excludes codes that
    are duplicated anywhere in the master -- unusable, so never offered."""
    candidates = index.rows_for(category, supplier)
    master = sorted((r for r in candidates
                     if r["unit_price"] and r["unit_price"] > 0),
                    key=lambda r: r["material_code"])
    if not master:
        raise ValueError(f"no usable items for {category} / {supplier}")
    n = len(master)
    # Only the duplicated codes that belong on THIS template. The old count
    # was every duplicate in the master, so a supplier with no duplicates at
    # all was still told that items were unavailable.
    excluded = index.excluded_for(category, supplier)
    no_price = len(candidates) - n

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"CML purchase order - request template ({CATEGORY_LABEL[category]})"
    t.font = Font(name=FONT, size=14, bold=True, color="187B85")
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    for i, text in enumerate((
            "Fill in only the shaded cells: Material Code, Qty and Note. "
            "Item, Supplier and Pack fill in by themselves.",
            f"Maximum {MAX_LINES} line items per order. Need more? Send a second "
            "file - each file becomes its own PO.",
            "Do not add or delete rows, rename the tabs, or edit the Master tab. "
            "Prices are not shown here and are not needed.")):
        r = 2 + i
        ws.merge_cells(f"A{r}:G{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = Font(name=FONT, size=9, color=MUTED)
        c.alignment = Alignment(vertical="center")

    for ref, label in (("A5", "Template version"), ("C5", "Generated"),
                       ("E5", "List")):
        ws[ref] = label
        ws[ref].font = Font(name=FONT, size=9, bold=True, color=MUTED)
    ws[META_CELLS["version"]] = TEMPLATE_VERSION
    ws[META_CELLS["generated"]] = local_now().strftime("%d-%b-%Y %H:%M")
    ws[META_CELLS["category"]] = CATEGORY_LABEL[category]
    ws[META_CELLS["supplier"]] = supplier
    for ref in META_CELLS.values():
        ws[ref].font = Font(name=FONT, size=9, color=MUTED)

    for i, (head, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=7, column=i, value=head)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX
    ws.row_dimensions[7].height = 20

    for offset in range(MAX_LINES):
        r = DATA_FIRST + offset
        ws.row_dimensions[r].height = 17
        no = ws.cell(row=r, column=1, value=offset + 1)
        no.font = Font(name=FONT, size=10, color=MUTED)
        no.alignment = Alignment(horizontal="center")
        no.border = BOX
        for col in (COL_CODE, COL_QTY, COL_NOTE):
            c = ws.cell(row=r, column=col)
            c.font = Font(name=FONT, size=10, color=INPUT_BLUE)
            c.fill = INPUT_FILL
            c.border = BOX
            c.protection = Protection(locked=False)
            if col == COL_QTY:
                c.alignment = Alignment(horizontal="center")
                c.number_format = "0"
        for col_i, master_col in ((3, "B"), (4, "C"), (5, "D")):
            f = ws.cell(row=r, column=col_i,
                        value=_lookup(f"$B{r}", master_col, n))
            f.font = Font(name=FONT, size=10, color=INK)
            f.fill = LOOKUP_FILL
            f.border = BOX

    guard = DATA_LAST + 1
    ws.merge_cells(f"A{guard}:G{guard}")
    g = ws[f"A{guard}"]
    g.value = (f"Anything typed below this line is ignored. Only rows "
               f"{DATA_FIRST}-{DATA_LAST} are read.")
    g.font = Font(name=FONT, size=9, italic=True, color="A32D2D")

    dv = DataValidation(type="list", formula1=f"={MASTER_SHEET}!$A$2:$A${n + 1}",
                        allow_blank=True, showErrorMessage=True,
                        errorTitle="Unknown material code",
                        error="Pick a code from the list. If the item you need is "
                              "missing, ask the bot for a fresh template.")
    ws.add_data_validation(dv)
    dv.add(f"B{DATA_FIRST}:B{DATA_LAST}")
    dvq = DataValidation(type="whole", operator="between", formula1=1,
                         formula2=9999, allow_blank=True, showErrorMessage=True,
                         errorTitle="Invalid quantity",
                         error="Quantity must be a whole number of 1 or more.")
    ws.add_data_validation(dvq)
    dvq.add(f"F{DATA_FIRST}:F{DATA_LAST}")

    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.freeze_panes = f"A{DATA_FIRST}"

    ms = wb.create_sheet(MASTER_SHEET)
    for i, head in enumerate(["Material Code", "Item", "Supplier", "Pack"], 1):
        c = ms.cell(row=1, column=i, value=head)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
    for i, w in enumerate([20, 44, 26, 16], 1):
        ms.column_dimensions[get_column_letter(i)].width = w
    for r, row in enumerate(master, start=2):
        for col, key in enumerate(("raw_code", "item", "supplier", "pack"), 1):
            c = ms.cell(row=r, column=col, value=row[key])
            c.font = Font(name=FONT, size=10, color=INK)
    ms.cell(row=1, column=6, value="Reference copy. No prices. Do not edit.").font = \
        Font(name=FONT, size=9, italic=True, color=MUTED)
    ms.protection.sheet = True
    ms.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(out_path or buf)
    return {"path": out_path, "bytes": None if out_path else buf.getvalue(),
            "n_items": n, "n_excluded": excluded, "n_no_price": no_price,
            "filename": _template_name(category, supplier)}


def _template_name(category, supplier):
    slug = re.sub(r"[^A-Za-z0-9]+", "_", supplier).strip("_")[:40]
    return f"PO_{category}_{slug}.xlsx"


class ReadError(Exception):
    pass


def read_request(data):
    """data: bytes of the uploaded .xlsx. Returns (rows, meta).

    rows always has exactly MAX_LINES entries -- one per cell of the fixed data
    range, blanks included -- so rows_read is a constant the completeness check
    can rely on. meta carries the workbook-level counts taken BEFORE row
    iteration, which is the only evidence that no row was silently skipped.
    """
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=False)
    except Exception as e:
        raise ReadError(f"I could not open that file as an Excel workbook ({e}).")
    if SHEET_NAME not in wb.sheetnames:
        raise ReadError(
            f"That workbook has no \u201c{SHEET_NAME}\u201d sheet. Please use the "
            f"template I sent you.")
    ws = wb[SHEET_NAME]

    rows = []
    for r in range(DATA_FIRST, DATA_LAST + 1):
        rows.append({"row_no": r,
                     "code": ws.cell(row=r, column=COL_CODE).value,
                     "qty": ws.cell(row=r, column=COL_QTY).value,
                     "note": ws.cell(row=r, column=COL_NOTE).value})

    below = 0
    for r in range(DATA_LAST + 2, min(ws.max_row, DATA_LAST + 60) + 1):
        code = ws.cell(row=r, column=COL_CODE).value
        qty = ws.cell(row=r, column=COL_QTY).value
        if norm_code(code) and str(qty or "").strip():
            below += 1

    cat_raw = str(ws[META_CELLS["category"]].value or "").strip().casefold()
    category = None
    for code, label in CATEGORY_LABEL.items():
        if cat_raw in (code, label.casefold()):
            category = code
    meta = {
        "sha256": hashlib.sha256(data).hexdigest(),
        "size": len(data),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": ",".join(wb.sheetnames),
        "max_row": ws.max_row,
        "rows_in_range": len(rows),
        "populated_below_range": below,
        "template_version": str(ws[META_CELLS["version"]].value or "").strip(),
        "template_generated": str(ws[META_CELLS["generated"]].value or "").strip(),
        "template_supplier": str(ws[META_CELLS["supplier"]].value or "").strip(),
        "template_category": category,
    }
    wb.close()
    return rows, meta


REPORT_HEADERS = ["Row", "Material Code", "Qty", "Note", "Status", "What to do"]
REPORT_WIDTHS = [7, 22, 9, 30, 18, 72]


def write_report(result, meta, out_path=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation report"
    s = result.summary

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "CML purchase order - upload validation report"
    t.font = Font(name=FONT, size=14, bold=True, color="187B85")

    lines = [
        f"Checked {local_now().strftime('%d-%b-%Y %H:%M')}",
        f"{s['rows_ok']} ready, {s['rows_blocked']} need fixing "
        f"({s['rows_populated']} lines used of {MAX_LINES})",
        "Nothing has been submitted. Fix the rows marked below in your own file "
        "and send it to me again.",
    ]
    if meta.get("populated_below_range"):
        lines.append(f"Warning: {meta['populated_below_range']} row(s) below the "
                     f"marked line were ignored. Move them into rows "
                     f"{DATA_FIRST}-{DATA_LAST}.")
    for i, text in enumerate(lines):
        r = 2 + i
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = Font(name=FONT, size=10,
                      color="A32D2D" if text.startswith("Warning") else MUTED)

    head = 3 + len(lines)
    for i, (h, w) in enumerate(zip(REPORT_HEADERS, REPORT_WIDTHS), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=head, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.border = BOX
    ws.freeze_panes = f"A{head + 1}"

    for i, row in enumerate(result.report, start=1):
        r = head + i
        ok = row["status"] == STATUS_OK
        values = [row["row_no"], row["raw_code"], row["raw_qty"], row["raw_note"],
                  "OK" if ok else row["status"].replace("_", " "),
                  row["message"] or row["matched_item"]]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name=FONT, size=10, color=INK)
            c.fill = OK_FILL if ok else BAD_FILL
            c.border = BOX
            c.alignment = Alignment(vertical="center", wrap_text=(col == 6))

    buf = BytesIO()
    wb.save(out_path or buf)
    return {"path": out_path, "bytes": None if out_path else buf.getvalue(),
            "filename": "PO_upload_report.xlsx"}
