"""The three small Excel side-files: stock count, price confirmation, and the
monthly outstanding/cancellation review.

Same shape as the other Excel layers -- openpyxl only, no Telegram, no gspread.
Each has a build_* that generates a locked file with one editable column, a
read_* that returns raw rows, and a validate_* that is a pure function.

The locked context columns are never trusted on the way back in: line_id,
ordered quantity and master price are all re-read from the sheet, so an edited
cell cannot change an outcome.
"""
import re
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

from clock import local_now

DATA_FIRST = 7
FONT = "Arial"
INK = "1F1F1F"
MUTED = "5F5E5A"
BLUE = "0000FF"
HEAD_FILL = PatternFill("solid", fgColor="187B85")
INPUT_FILL = PatternFill("solid", fgColor="FFF6D9")
LOCK_FILL = PatternFill("solid", fgColor="F4F3EF")
HAIR = Side(style="thin", color="C9C7BF")
BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

NA = {"#n/a", "n/a", "na", "-"}


def _norm(v):
    """Text of a cell, with 0 preserved.

    `str(v or "")` turned an Excel integer 0 into an empty string, because 0 is
    falsy. The stock controller typing 0 -- none on the shelf, the strongest
    argument there is FOR the order -- was told "on-hand is required".
    """
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _to_float(v):
    if v is None or isinstance(v, bool):
        return None
    s = str(v).replace(",", "").replace("$", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(v):
    f = _to_float(v)
    if f is None or f != int(f) or f < 0:
        return None
    return int(f)


def _sheet(wb, title, headers, widths, note):
    ws = wb.active
    ws.title = title
    span = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{span}1")
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=14, bold=True, color="187B85")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(f"A2:{span}2")
    ws["A2"] = note
    ws["A2"].font = Font(name=FONT, size=9, color=MUTED)
    ws.merge_cells(f"A3:{span}3")
    ws["A3"] = f"Generated {local_now():%d-%b-%Y %H:%M}"
    ws["A3"].font = Font(name=FONT, size=9, italic=True, color=MUTED)
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=DATA_FIRST - 1, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BOX
    ws.row_dimensions[DATA_FIRST - 1].height = 26
    return ws


def _cell(ws, row, col, value, editable=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT, size=10, color=BLUE if editable else INK)
    c.fill = INPUT_FILL if editable else LOCK_FILL
    c.border = BOX
    if editable:
        c.protection = Protection(locked=False)
    if fmt:
        c.number_format = fmt
    return c


def _finish(wb, ws, meta):
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.freeze_panes = f"A{DATA_FIRST}"
    mw = wb.create_sheet("_meta")
    for i, (k, v) in enumerate(meta.items(), start=1):
        mw.cell(row=i, column=1, value=k)
        mw.cell(row=i, column=2, value=str(v))
    mw.sheet_state = "hidden"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read(data, title, cols, meta_keys):
    meta = {"error": "", "kind": "", "po_no": "", "data_last": 0}
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:
        meta["error"] = f"not a readable Excel file ({type(exc).__name__})"
        return [], meta
    if "_meta" in wb.sheetnames:
        mw = wb["_meta"]
        got = {str(mw.cell(row=i, column=1).value or ""):
               str(mw.cell(row=i, column=2).value or "")
               for i in range(1, mw.max_row + 1)}
        for k in meta_keys:
            meta[k] = got.get(k, "")
    if title not in wb.sheetnames:
        meta["error"] = f"no '{title}' tab in this file. Ask for a fresh one."
        return [], meta
    ws = wb[title]
    try:
        last = int(meta.get("data_last") or ws.max_row)
    except ValueError:
        last = ws.max_row
    rows = []
    for r in range(DATA_FIRST, last + 1):
        row = {"row_no": r}
        for key, col in cols.items():
            row[key] = ws.cell(row=r, column=col).value
        if any(_norm(v) for k, v in row.items() if k != "row_no"):
            rows.append(row)
    meta["data_last"] = last
    return rows, meta


# ===================== stock count =====================

STOCK_TITLE = "Stock count"
STOCK_HEADERS = ["No.", "CML Code", "Item", "Pack", "Requested", "On hand"]
STOCK_COLS = {"code": 2, "on_hand": 6}


def build_stock_file(po_no, lines):
    wb = Workbook()
    ws = _sheet(wb, STOCK_TITLE, STOCK_HEADERS, [6, 16, 48, 18, 12, 14],
                "Enter the units you currently hold for each line. "
                "Enter #N/A if a count is not possible. Do not add or delete rows.")
    r = DATA_FIRST
    for n, it in enumerate(lines, start=1):
        _cell(ws, r, 1, n)
        _cell(ws, r, 2, it.get("material_code", "")).number_format = "@"
        _cell(ws, r, 3, it.get("item", ""))
        _cell(ws, r, 4, it.get("pack", ""))
        _cell(ws, r, 5, it.get("qty", ""))
        _cell(ws, r, 6, None, editable=True)
        r += 1
    return {"bytes": _finish(wb, ws, {"kind": "stock", "po_no": po_no,
                                      "data_last": r - 1}),
            "filename": f"Stock_count_PO_{po_no}.xlsx", "lines": len(lines)}


def read_stock(data):
    return _read(data, STOCK_TITLE, STOCK_COLS, ("kind", "po_no", "data_last"))


def validate_stock(rows, lines):
    """Mandatory means every line answered. #N/A is a permitted answer -- and
    it must never coerce to 0, because 0 means 'none in stock', the strongest
    possible argument FOR the order."""
    by_code = {_norm(l.get("material_code")).upper(): l for l in lines}
    counts, errs, seen, attempted = [], [], set(), set()
    for row in rows:
        code = _norm(row.get("code")).upper()
        line = by_code.get(code)
        if not line:
            errs.append({"row_no": row["row_no"],
                         "message": f"{code or '(blank)'} is not on this PO."})
            continue
        attempted.add(code)
        raw = _norm(row.get("on_hand"))
        if not raw:
            errs.append({"row_no": row["row_no"],
                         "message": f"{line['item']}: on-hand is required. "
                                    f"Enter #N/A if you cannot count it."})
            continue
        if raw.casefold() in NA:
            val = "#N/A"
        else:
            n = _to_int(raw)
            if n is None:
                errs.append({"row_no": row["row_no"],
                             "message": f"{line['item']}: on-hand must be a "
                                        f"whole number of 0 or more, or #N/A."})
                continue
            val = n
        seen.add(code)
        counts.append({"line_id": line["line_id"], "code": code,
                       "item": line["item"], "on_hand": val,
                       "ordered": line.get("qty", "")})
    # Only lines the file never mentioned at all -- a line that was mentioned
    # but left blank already has its own, more specific error.
    missing = [l for c, l in by_code.items() if c not in seen | attempted]
    for l in missing:
        errs.append({"row_no": "-",
                     "message": f"{l['item']}: no on-hand entered."})
    return counts, errs


# ===================== price confirmation =====================

PRICE_TITLE = "Price confirmation"
PRICE_HEADERS = ["No.", "CML Code", "Supplier code", "Item", "Qty",
                 "Master price", "Confirmed price"]
PRICE_COLS = {"code": 2, "new_price": 7}


def build_price_file(po_no, lines):
    wb = Workbook()
    ws = _sheet(wb, PRICE_TITLE, PRICE_HEADERS, [6, 16, 18, 44, 9, 14, 16],
                "Enter the price the supplier confirmed. Leave a line blank to "
                "keep the master-list price. Any change is flagged to Finance, "
                "GM and the Board.")
    r = DATA_FIRST
    for n, it in enumerate(lines, start=1):
        _cell(ws, r, 1, n)
        _cell(ws, r, 2, it.get("material_code", "")).number_format = "@"
        _cell(ws, r, 3, it.get("supplier_code", "")).number_format = "@"
        _cell(ws, r, 4, it.get("item", ""))
        _cell(ws, r, 5, it.get("qty", ""))
        _cell(ws, r, 6, it.get("unit_price", ""), fmt="#,##0.00")
        _cell(ws, r, 7, None, editable=True, fmt="#,##0.00")
        r += 1
    return {"bytes": _finish(wb, ws, {"kind": "price", "po_no": po_no,
                                      "data_last": r - 1}),
            "filename": f"Price_confirm_PO_{po_no}.xlsx", "lines": len(lines)}


def read_price(data):
    return _read(data, PRICE_TITLE, PRICE_COLS, ("kind", "po_no", "data_last"))


def validate_price(rows, lines):
    by_code = {_norm(l.get("material_code")).upper(): l for l in lines}
    changes, errs = [], []
    for row in rows:
        code = _norm(row.get("code")).upper()
        line = by_code.get(code)
        if not line:
            errs.append({"row_no": row["row_no"],
                         "message": f"{code or '(blank)'} is not on this PO."})
            continue
        raw = _norm(row.get("new_price"))
        if not raw:
            continue
        val = _to_float(raw)
        if val is None or val <= 0:
            errs.append({"row_no": row["row_no"],
                         "message": f"{line['item']}: price must be a positive "
                                    f"number."})
            continue
        old = float(line.get("unit_price") or 0)
        if abs(val - old) <= 0.005:
            continue
        qty = int(line.get("qty") or 0)
        changes.append({"line_id": line["line_id"], "item": line["item"],
                        "old_price": old, "new_price": round(val, 2),
                        "new_total": round(val * qty, 2)})
    return changes, errs


# ===================== outstanding / cancellation =====================

OUT_TITLE = "Outstanding lines"
OUT_HEADERS = ["PO", "Line", "CML Code", "Item", "Supplier", "Raised",
               "Ordered", "Received", "Outstanding", "Remove?", "Reason"]
OUT_COLS = {"po_no": 1, "line_id": 2, "remove": 10, "reason": 11}


def build_outstanding_file(rows):
    wb = Workbook()
    ws = _sheet(wb, OUT_TITLE, OUT_HEADERS,
                [8, 12, 16, 42, 22, 14, 10, 10, 12, 11, 34],
                "Put x in Remove and give a reason to cancel the un-received "
                "remainder of a line. Received quantities are never touched, "
                "and lines are marked cancelled rather than deleted.")
    r = DATA_FIRST
    for it in rows:
        _cell(ws, r, 1, it["po_no"])
        _cell(ws, r, 2, it["line_id"])
        _cell(ws, r, 3, it["material_code"]).number_format = "@"
        _cell(ws, r, 4, it["item"])
        _cell(ws, r, 5, it["supplier"])
        _cell(ws, r, 6, it["created_at"])
        _cell(ws, r, 7, it["ordered"])
        _cell(ws, r, 8, it["received"])
        _cell(ws, r, 9, it["outstanding"])
        _cell(ws, r, 10, None, editable=True)
        _cell(ws, r, 11, None, editable=True)
        r += 1
    return {"bytes": _finish(wb, ws, {"kind": "cancel", "data_last": r - 1}),
            "filename": f"Outstanding_{local_now():%Y%m%d}.xlsx",
            "lines": len(rows)}


def read_outstanding(data):
    return _read(data, OUT_TITLE, OUT_COLS, ("kind", "data_last"))


def validate_cancel(rows, open_lines):
    """Cancels only the UN-RECEIVED remainder. A line at 6-of-10 cancels 4 and
    stays at 6 received; a fully-received line cannot be cancelled at all."""
    by_id = {(str(l["po_no"]), str(l["line_id"])): l for l in open_lines}
    removals, errs = [], []
    for row in rows:
        mark = _norm(row.get("remove")).casefold()
        if mark not in ("x", "yes", "y", "1", "true"):
            continue
        key = (_norm(row.get("po_no")), _norm(row.get("line_id")))
        line = by_id.get(key)
        if not line:
            errs.append({"row_no": row["row_no"],
                         "message": f"PO #{key[0]} line {key[1]} is no longer "
                                    f"outstanding \u2014 it may have been "
                                    f"received since this file was made."})
            continue
        reason = _norm(row.get("reason"))
        if not reason:
            errs.append({"row_no": row["row_no"],
                         "message": f"{line['item']}: a reason is required."})
            continue
        if line["outstanding"] <= 0:
            errs.append({"row_no": row["row_no"],
                         "message": f"{line['item']}: nothing outstanding to "
                                    f"cancel."})
            continue
        removals.append({"po_no": line["po_no"], "line_id": line["line_id"],
                         "item": line["item"], "qty": line["outstanding"],
                         "reason": reason})
    return removals, errs
